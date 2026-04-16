#!/usr/bin/env python3
"""
NUC Accreditation Import Script
Imports 11,055 programs with full accreditation history from Excel file
Usage: python3 import-accreditation.py
"""

import os
import re
import psycopg2
from openpyxl import load_workbook

DATABASE_URL = os.environ.get('DATABASE_URL', '')
if not DATABASE_URL:
    print("❌ Set DATABASE_URL environment variable first")
    exit(1)

YEAR_COLS = {
    8: 1990, 9: 1999, 10: 2002, 11: 2005, 12: 2006,
    13: 2007, 14: 2008, 15: 2009, 16: 2010, 17: 2011,
    18: 2012, 19: 2013, 20: 2014, 21: 2015, 22: 2016,
    23: 2017, 24: 2018, 25: 2019, 26: 2020, 27: 2021,
    28: 2022, 29: 2023, 30: 2024
}

STATUS_MAP = {
    'full': 'FULL', 'fulll': 'FULL',
    'interim': 'INTERIM', 'interrim': 'INTERIM', 'interin': 'INTERIM',
    'denied': 'DENIED', 'denied (interim b/b)': 'DENIED',
    'suspended': 'DENIED', 'with-held': 'DENIED',
}

DEGREE_MAP = {
    'b.sc.': 'BSC', 'b.sc': 'BSC', 'bsc': 'BSC',
    'b.a.': 'BA', 'b.a': 'BA', 'ba': 'BA',
    'b. eng.': 'BENG', 'b.eng': 'BENG',
    'b. tech.': 'BTECH', 'b.tech': 'BTECH',
    'll.b.': 'LLB', 'll. b.': 'LLB',
    'mbbs': 'MBBS', 'mb.bs': 'MBBS',
    'b. pharm.': 'BPHARM', 'pharm d.': 'BPHARM',
    'b. agric.': 'BSC', 'b. mls': 'BSC',
    'b. nsc.': 'BSC', 'o. d.': 'BSC',
    'dvm': 'BSC', 'dpt': 'BSC', 'blis': 'BSC',
    'b. lis': 'BSC', 'b. ed.': 'BA', 'b.sc. (ed.)': 'BSC',
    'b.a. (ed.)': 'BA', 'bhim': 'BSC',
}


def make_slug(text):
    import re
    s = text.lower().strip()
    s = re.sub(r'[^\w\s-]', '', s)
    s = re.sub(r'[\s_-]+', '-', s)
    s = re.sub(r'^-+|-+$', '', s)
    return s[:80]

def normalize_status(val):
    if not val:
        return None
    v = str(val).strip().lower()
    return STATUS_MAP.get(v, None)

def normalize_degree(val):
    if not val:
        return 'BSC'
    v = str(val).strip().lower()
    return DEGREE_MAP.get(v, 'BSC')

def main():
    print("🔌 Connecting to database...")
    conn = psycopg2.connect(DATABASE_URL)
    cur = conn.cursor()

    print("📖 Reading Excel file...")
    wb = load_workbook(
        "/Users/sainteni/Downloads/Approved_Programmes_and_Accreditation_Results as at 2025 (2) (1).xlsx",
        read_only=True
    )
    ws = wb.active

    # Cache universities
    cur.execute('SELECT id, name, slug FROM "University"')
    unis = cur.fetchall()
    uni_map = {}
    for uid, uname, uslug in unis:
        uni_map[uname.lower().strip()] = uid
        uni_map[uslug.lower().strip()] = uid

    # Cache faculties
    cur.execute('SELECT id, name FROM "Faculty"')
    fac_rows = cur.fetchall()
    fac_map = {name.lower().strip(): fid for fid, name in fac_rows}

    created_programs = 0
    created_accreditations = 0
    skipped = 0
    errors = 0
    slugs_seen = set()

    print("🚀 Importing programs and accreditation history...\n")

    rows_processed = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1] or not row[3]:
            continue

        uni_name = str(row[1]).strip()
        discipline = str(row[2]).strip() if row[2] else ''
        program_name = str(row[3]).strip()
        degree_raw = str(row[6]).strip() if row[6] else 'B.Sc.'
        year_est = int(row[7]) if row[7] and str(row[7]).isdigit() else None
        maturity_raw = row[31]
        maturity_year = int(maturity_raw) if maturity_raw and str(maturity_raw).replace('.0','').isdigit() else None

        # Find university
        uni_id = None
        uni_lower = uni_name.lower()
        # Try exact match first
        uni_id = uni_map.get(uni_lower)
        if not uni_id:
            # Try partial match
            for key, uid in uni_map.items():
                if len(key) > 5 and (key in uni_lower or uni_lower in key):
                    uni_id = uid
                    break

        if not uni_id:
            skipped += 1
            continue

        # Get or create faculty
        fac_id = None
        if discipline:
            fac_lower = discipline.lower()
            fac_id = fac_map.get(fac_lower)
            if not fac_id:
                try:
                    fac_slug = make_slug(discipline)
                    cur.execute(
                        'INSERT INTO "Faculty" (id, name, slug, "createdAt", "updatedAt") VALUES (gen_random_uuid(), %s, %s, NOW(), NOW()) ON CONFLICT (slug) DO UPDATE SET name = EXCLUDED.name RETURNING id',
                        (discipline, fac_slug)
                    )
                    fac_id = cur.fetchone()[0]
                    fac_map[fac_lower] = fac_id
                    conn.commit()
                except Exception as e:
                    conn.rollback()

        # Create program slug
        prog_slug = make_slug(f"{program_name}-{str(uni_id)[:8]}")
        orig_slug = prog_slug
        counter = 1
        while prog_slug in slugs_seen:
            prog_slug = f"{orig_slug}-{counter}"
            counter += 1
        slugs_seen.add(prog_slug)

        degree = normalize_degree(degree_raw)

        # Create program
        try:
            cur.execute(
                '''INSERT INTO "Program" (id, name, slug, "universityId", "facultyId", "degreeType", "isActive", "createdAt", "updatedAt")
                   VALUES (gen_random_uuid(), %s, %s, %s, %s, %s, true, NOW(), NOW())
                   ON CONFLICT DO NOTHING RETURNING id''',
                (program_name, prog_slug, uni_id, fac_id, degree)
            )
            result = cur.fetchone()
            if not result:
                # Try to find existing
                cur.execute('SELECT id FROM "Program" WHERE "universityId" = %s AND slug = %s', (uni_id, prog_slug))
                result = cur.fetchone()

            if not result:
                skipped += 1
                conn.rollback()
                continue

            prog_id = result[0]
            conn.commit()
            created_programs += 1

            # First mark all previous as not current
            cur.execute('UPDATE "Accreditation" SET "isCurrent" = false WHERE "programId" = %s', (prog_id,))

            # Insert accreditation history for each year
            latest_status = None
            latest_year = None
            for col_idx, year in YEAR_COLS.items():
                if col_idx < len(row):
                    status = normalize_status(row[col_idx])
                    if status:
                        cur.execute(
                            '''INSERT INTO "Accreditation" (id, "programId", status, year, "isCurrent", "createdAt", "updatedAt")
                               VALUES (gen_random_uuid(), %s, %s, %s, false, NOW(), NOW())''',
                            (prog_id, status, year)
                        )
                        latest_status = status
                        latest_year = year
                        created_accreditations += 1

            # Mark latest as current with maturity date
            if latest_status and latest_year:
                import datetime
                expiry = datetime.date(maturity_year, 12, 31) if maturity_year else None
                cur.execute(
                    '''INSERT INTO "Accreditation" (id, "programId", status, year, "expiryDate", "isCurrent", "createdAt", "updatedAt")
                       VALUES (gen_random_uuid(), %s, %s, %s, %s, true, NOW(), NOW())''',
                    (prog_id, latest_status, latest_year, expiry)
                )
                created_accreditations += 1

            conn.commit()

        except Exception as e:
            errors += 1
            conn.rollback()
            if errors <= 3:
                print(f"  ✗ {program_name} @ {uni_name[:30]}: {e}")

        rows_processed += 1
        if rows_processed % 500 == 0:
            print(f"  ✓ Processed {rows_processed} rows... ({created_programs} programs, {created_accreditations} accreditation records)")

    # Update stats
    cur.execute('SELECT COUNT(*) FROM "Program" WHERE "isActive" = true')
    total_progs = cur.fetchone()[0]
    cur.execute('SELECT COUNT(*) FROM "Accreditation" WHERE "isCurrent" = true AND status = \'FULL\'')
    accredited = cur.fetchone()[0]
    cur.execute(
        '''UPDATE "PlatformStat" SET "totalPrograms" = %s, "accreditedPrograms" = %s, "updatedAt" = NOW() WHERE id = 'singleton' ''',
        (total_progs, accredited)
    )
    conn.commit()

    cur.close()
    conn.close()

    print(f"\n✅ Import complete!")
    print(f"   Programs created: {created_programs}")
    print(f"   Accreditation records: {created_accreditations}")
    print(f"   Skipped (no matching university): {skipped}")
    print(f"   Errors: {errors}")
    print(f"   Total programs in DB: {total_progs}")
    print(f"   Fully accredited: {accredited}")

if __name__ == '__main__':
    main()
