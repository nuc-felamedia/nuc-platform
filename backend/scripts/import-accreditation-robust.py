#!/usr/bin/env python3
import os, re, sys, json, datetime, time, psycopg2
from openpyxl import load_workbook

DATABASE_URL = os.environ.get('DATABASE_URL', '')
EXCEL_FILE = os.path.expanduser(
    "~/Downloads/Approved_Programmes_and_Accreditation_Results as at 2025 (2) (1).xlsx"
)

YEAR_COLS = {8:1990,9:1999,10:2002,11:2005,12:2006,13:2007,14:2008,15:2009,
             16:2010,17:2011,18:2012,19:2013,20:2014,21:2015,22:2016,23:2017,
             24:2018,25:2019,26:2020,27:2021,28:2022,29:2023,30:2024}

STATUS_MAP = {'full':'FULL','fulll':'FULL','interim':'INTERIM','interrim':'INTERIM',
              'interin':'INTERIM','denied':'DENIED','denied (interim b/b)':'DENIED',
              'suspended':'DENIED','with-held':'DENIED'}

DEGREE_MAP = {'b.sc.':'BSC','b.sc':'BSC','b.a.':'BA','b.a':'BA','b. eng.':'BENG',
              'b. tech.':'BTECH','ll.b.':'LLB','ll. b.':'LLB','mbbs':'MBBS',
              'b. pharm.':'BPHARM','pharm d.':'BPHARM','b. agric.':'BSC',
              'b. mls':'BSC','b. nsc.':'BSC','o. d.':'BSC','dvm':'BSC','dpt':'BSC',
              'blis':'BSC','b. lis':'BSC','b. ed.':'BA','b.sc. (ed.)':'BSC',
              'b.a. (ed.)':'BA','bhim':'BSC','b.a':'BA'}

def normalize_status(val):
    if not val: return None
    return STATUS_MAP.get(str(val).strip().lower(), None)

def normalize_degree(val):
    if not val: return 'BSC'
    return DEGREE_MAP.get(str(val).strip().lower(), 'BSC')

def make_slug(text):
    s = str(text).lower().strip()
    s = re.sub(r'[^\w\s-]', '', s)
    s = re.sub(r'[\s_]+', '-', s)
    s = re.sub(r'-+', '-', s).strip('-')
    return s[:80]

def get_connection():
    for attempt in range(5):
        try:
            conn = psycopg2.connect(DATABASE_URL, connect_timeout=30)
            conn.autocommit = False
            return conn
        except Exception as e:
            print(f"  Connection attempt {attempt+1} failed: {e}")
            time.sleep(5 * (attempt + 1))
    raise Exception("Could not connect after 5 attempts")

def main():
    print("🔌 Connecting to database...")
    conn = get_connection()
    cur = conn.cursor()
    print("✅ Connected\n")

    print("📖 Reading Excel file...")
    wb = load_workbook(EXCEL_FILE, read_only=True)
    ws = wb.active
    print("✅ File loaded\n")

    cur.execute('SELECT LOWER(name), "universityId" FROM "Program"')
    existing_programs = set(cur.fetchall())
    print(f"   Skipping {len(existing_programs)} already-imported programs\n")

    cur.execute('SELECT id, LOWER(name), slug FROM "University"')
    uni_rows = cur.fetchall()
    uni_map = {name: uid for uid, name, slug in uni_rows}

    cur.execute('SELECT id, LOWER(name) FROM "Faculty"')
    fac_map = {name: fid for fid, name in cur.fetchall()}

    cur.execute('SELECT slug FROM "Program"')
    slugs_seen = set(s for (s,) in cur.fetchall())

    created = 0; skipped = 0; errors = 0; rows_done = 0

    print("🚀 Importing...\n")

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1] or not row[3]:
            continue

        uni_name = str(row[1]).strip()
        discipline = str(row[2]).strip() if row[2] else ''
        prog_name = str(row[3]).strip()

        uni_id = uni_map.get(uni_name.lower())
        if not uni_id:
            uni_lower = uni_name.lower()
            for key, uid in uni_map.items():
                if len(key) > 8 and (key in uni_lower or uni_lower in key):
                    uni_id = uid
                    break

        if not uni_id:
            skipped += 1
            rows_done += 1
            continue

        if (prog_name.lower(), uni_id) in existing_programs:
            skipped += 1
            rows_done += 1
            continue

        fac_id = None
        if discipline:
            fac_id = fac_map.get(discipline.lower())
            if not fac_id:
                try:
                    fac_slug = make_slug(discipline)
                    cur.execute(
                        'INSERT INTO "Faculty" (id, name, slug, "createdAt", "updatedAt") VALUES (gen_random_uuid(), %s, %s, NOW(), NOW()) ON CONFLICT (slug) DO UPDATE SET name = EXCLUDED.name RETURNING id',
                        (discipline, fac_slug)
                    )
                    fac_id = cur.fetchone()[0]
                    fac_map[discipline.lower()] = fac_id
                    conn.commit()
                except Exception:
                    conn.rollback()

        base_slug = make_slug(f"{prog_name}-{str(uni_id)[:8]}")
        prog_slug = base_slug
        counter = 1
        while prog_slug in slugs_seen:
            prog_slug = f"{base_slug}-{counter}"
            counter += 1
        slugs_seen.add(prog_slug)

        degree = normalize_degree(str(row[6]).strip() if row[6] else '')
        maturity_raw = row[31]
        maturity_year = None
        if maturity_raw:
            try: maturity_year = int(float(str(maturity_raw)))
            except: pass

        try:
            cur.execute(
                'INSERT INTO "Program" (id, name, slug, "universityId", "facultyId", "degreeType", "isActive", "createdAt", "updatedAt") VALUES (gen_random_uuid(), %s, %s, %s, %s, %s, true, NOW(), NOW()) ON CONFLICT DO NOTHING RETURNING id',
                (prog_name, prog_slug, uni_id, fac_id, degree)
            )
            result = cur.fetchone()
            if not result:
                skipped += 1
                conn.rollback()
                rows_done += 1
                continue

            prog_id = result[0]
            latest_status = None
            latest_year = None

            for col_idx, year in YEAR_COLS.items():
                if col_idx < len(row):
                    status = normalize_status(row[col_idx])
                    if status:
                        cur.execute(
                            'INSERT INTO "Accreditation" (id, "programId", status, year, "isCurrent", "createdAt", "updatedAt") VALUES (gen_random_uuid(), %s, %s, %s, false, NOW(), NOW())',
                            (prog_id, status, year)
                        )
                        latest_status = status
                        latest_year = year

            if latest_status and latest_year:
                expiry = datetime.date(maturity_year, 12, 31) if maturity_year else None
                cur.execute(
                    'INSERT INTO "Accreditation" (id, "programId", status, year, "expiryDate", "isCurrent", "createdAt", "updatedAt") VALUES (gen_random_uuid(), %s, %s, %s, %s, true, NOW(), NOW())',
                    (prog_id, latest_status, latest_year, expiry)
                )

            conn.commit()
            existing_programs.add((prog_name.lower(), uni_id))
            created += 1

        except psycopg2.InterfaceError:
            print(f"\n  ⚠️  Connection dropped at row {rows_done}. Reconnecting...")
            try: conn.close()
            except: pass
            conn = get_connection()
            cur = conn.cursor()
            print(f"  ✅ Reconnected.\n")
        except Exception as e:
            errors += 1
            conn.rollback()
            if errors <= 3:
                print(f"  ✗ {prog_name[:40]}: {e}")

        rows_done += 1
        if rows_done % 500 == 0:
            print(f"  ✓ {rows_done} rows | {created} created | {skipped} skipped")

    cur.execute('SELECT COUNT(*) FROM "Program" WHERE "isActive" = true')
    total = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM \"Accreditation\" WHERE \"isCurrent\" = true AND status = 'FULL'")
    accredited = cur.fetchone()[0]
    cur.execute("UPDATE \"PlatformStat\" SET \"totalPrograms\" = %s, \"accreditedPrograms\" = %s, \"updatedAt\" = NOW() WHERE id = 'singleton'", (total, accredited))
    conn.commit()
    conn.close()

    print(f"\n✅ DONE — Created: {created} | Skipped: {skipped} | Errors: {errors}")
    print(f"   Total programs: {total:,} | Fully accredited: {accredited:,}")

if __name__ == '__main__':
    main()
